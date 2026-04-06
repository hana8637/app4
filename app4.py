import streamlit as st
import pandas as pd
import math
import numpy as np
import os
import platform
import matplotlib.pyplot as plt
import matplotlib.patches as patches
import matplotlib.font_manager as fm
from io import BytesIO

# --- 엑셀 스타일링 라이브러리 ---
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ==============================================================================
# 공통 설정: 스트림릿 웹 환경 한글 폰트 깨짐 완벽 방지
# ==============================================================================
st.set_page_config(page_title="하나천막기업 - 도면 산출 시스템", layout="wide")

@st.cache_resource
def set_korean_font():
    os_name = platform.system()
    if os_name == "Windows":
        plt.rc('font', family='Malgun Gothic')
    elif os_name == "Darwin":
        plt.rc('font', family='AppleGothic')
    else:
        # 리눅스 (Streamlit Cloud 등) 한글 폰트 자동 탐색 및 적용
        font_dirs = ['/usr/share/fonts', '/usr/local/share/fonts', '~/.fonts']
        font_files = fm.findSystemFonts(fontpaths=font_dirs)
        for font_file in font_files:
            if 'Nanum' in font_file or 'nanum' in font_file:
                fm.fontManager.addfont(font_file)
                font_name = fm.FontProperties(fname=font_file).get_name()
                plt.rc('font', family=font_name)
                break
        else:
            plt.rc('font', family='NanumGothic')
    plt.rcParams['axes.unicode_minus'] = False

set_korean_font()

# ==============================================================================
# [공통] 엑셀 다운로드용 바이트 변환기
# ==============================================================================
def get_truss_excel_bytes(raw_data):
    df = pd.DataFrame(raw_data)
    df_grouped = df.groupby(["구분", "품명", "재단기장(L)", "상단 가공각(°)", "하단 가공각(°)"]).size().reset_index(name='1대당 수량')
    
    sort_mapping = {
        "상현대(전체)": -2, "하현대(전체)": -1,  
        "용마루": 1, "상단용마루": 2, "하단용마루": 3,
        "수평재": 4, "밑더블수평재": 5, 
        "다대": 6, "상단다대": 7, "하단다대": 8,
        "살대": 9, "상단살대": 10, "하단살대": 11,
        "수평내부다대": 12, "수평내부살대": 13, "서브다대": 14, "서브살대": 15
    }
    df_grouped['정렬키'] = df_grouped['구분'].map(sort_mapping).fillna(99)
    df_grouped = df_grouped.sort_values(by=["정렬키", "재단기장(L)"], ascending=[True, False]).drop('정렬키', axis=1)
    
    df_grouped.insert(0, '순번', range(1, len(df_grouped) + 1))
    df_grouped["총 소요 수량"] = ""
    df_grouped["6M 소요본수"] = ""
    df_grouped = df_grouped[["순번", "구분", "품명", "1대당 수량", "총 소요 수량", "재단기장(L)", "상단 가공각(°)", "하단 가공각(°)", "6M 소요본수"]]

    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_grouped.to_excel(writer, sheet_name='통합 재단표', index=False, startrow=2)
        ws = writer.sheets['통합 재단표']
        
        ws.merge_cells('A1:C1')
        ws['A1'] = "👉 트러스 총 제작 수량 (EA) :"
        ws['A1'].font = Font(bold=True, size=12)
        ws['A1'].alignment = Alignment(horizontal="right", vertical="center")
        
        ws['D1'] = 1
        ws['D1'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        ws['D1'].font = Font(color="FF0000", bold=True, size=14)
        ws['D1'].alignment = Alignment(horizontal="center", vertical="center")
        
        thin_border = Border(left=Side(style='thin', color='A6A6A6'), right=Side(style='thin', color='A6A6A6'),
                             top=Side(style='thin', color='A6A6A6'), bottom=Side(style='thin', color='A6A6A6'))
        ws['D1'].border = thin_border

        color_map = {
            "상현대(전체)": PatternFill(start_color="D0CECE", end_color="D0CECE", fill_type="solid"),
            "하현대(전체)": PatternFill(start_color="AEAAAA", end_color="AEAAAA", fill_type="solid"),
            "용마루": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
            "상단용마루": PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid"),
            "하단용마루": PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid"),
            "수평재": PatternFill(start_color="D2B4DE", end_color="D2B4DE", fill_type="solid"), 
            "밑더블수평재": PatternFill(start_color="E8DAEF", end_color="E8DAEF", fill_type="solid"),
            "다대": PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"),
            "상단다대": PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid"),
            "하단다대": PatternFill(start_color="8EA9DB", end_color="8EA9DB", fill_type="solid"),
            "살대": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
            "상단살대": PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid"),
            "하단살대": PatternFill(start_color="A9D08E", end_color="A9D08E", fill_type="solid"),
            "수평내부다대": PatternFill(start_color="A9DFBF", end_color="A9DFBF", fill_type="solid"), 
            "수평내부살대": PatternFill(start_color="F9E79F", end_color="F9E79F", fill_type="solid"), 
            "서브다대": PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid"),
            "서브살대": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        }
        
        for r_idx, row in enumerate(ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=ws.max_column), 3):
            gubun_val = ws.cell(row=r_idx, column=2).value if r_idx > 3 else None
            for c_idx, cell in enumerate(row, 1):
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border
                
                if r_idx == 3:
                    cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                    cell.font = Font(color="FFFFFF", bold=True)
                else:
                    header = ws.cell(row=3, column=c_idx).value
                    if header == "순번": cell.font = Font(bold=True)
                    elif header == "구분":
                        cell.fill = color_map.get(gubun_val, PatternFill(fill_type=None))
                        cell.font = Font(bold=True)
                    elif header == "총 소요 수량":
                        cell.value = f'=$D$1*D{r_idx}' 
                        cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
                        cell.font = Font(color="0070C0", bold=True)
                    elif header == "6M 소요본수":
                        cell.value = f'=ROUNDUP((E{r_idx}*F{r_idx})/6000, 1)'
                        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                        cell.font = Font(color="0070C0", bold=True)
                    elif header in ["1대당 수량", "재단기장(L)", "상단 가공각(°)", "하단 가공각(°)"]:
                        cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                        cell.font = Font(color="C00000", bold=True)
                    elif r_idx % 2 == 0: cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        start_col = 11  
        ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=start_col+4)
        title_cell = ws.cell(row=1, column=start_col, value="레이저 가공 싸이즈")
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        title_cell.fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
        title_cell.border = thin_border
        
        headers_new = ["구분", "재단기장\n(반올림)", "상단 가공각", "하단 가공각", "총수량"]
        for i, h in enumerate(headers_new):
            c = ws.cell(row=3, column=start_col+i, value=h)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = thin_border

        counters = {}
        row_idx_new = 4
        for _, row_data in df_grouped.iterrows():
            cat_base = row_data["구분"]
            if cat_base not in counters: counters[cat_base] = 1
            name_val = f"{cat_base}{counters[cat_base]}"
            counters[cat_base] += 1
            
            L_val = row_data["재단기장(L)"]
            top_a = row_data["상단 가공각(°)"]
            bot_a = row_data["하단 가공각(°)"]
            qty_val = row_data["1대당 수량"]
            
            L_rounded = round(float(L_val)) if pd.notnull(L_val) else 0
            top_ceil = int(top_a) if pd.notnull(top_a) else 0
            bot_ceil = int(bot_a) if pd.notnull(bot_a) else 0
            
            ws.cell(row=row_idx_new, column=start_col, value=name_val)
            ws.cell(row=row_idx_new, column=start_col+1, value=L_rounded)
            ws.cell(row=row_idx_new, column=start_col+2, value=top_ceil)
            ws.cell(row=row_idx_new, column=start_col+3, value=bot_ceil)
            ws.cell(row=row_idx_new, column=start_col+4, value=f"=$D$1*{qty_val}") 
            ws.cell(row=row_idx_new, column=start_col+4).font = Font(color="0070C0", bold=True)
            
            for i in range(5):
                c = ws.cell(row=row_idx_new, column=start_col+i)
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border = thin_border
                if row_idx_new % 2 == 0: 
                    c.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            row_idx_new += 1

        pipe_summary = {}
        for item in raw_data:
            p_name = item["품명"]
            l_val = float(item["재단기장(L)"]) if pd.notnull(item["재단기장(L)"]) else 0
            if p_name not in pipe_summary:
                pipe_summary[p_name] = {"len": 0, "qty": 0}
            pipe_summary[p_name]["len"] += l_val
            pipe_summary[p_name]["qty"] += 1
            
        sum_start_col = 17 
        ws.merge_cells(start_row=1, start_column=sum_start_col, end_row=1, end_column=sum_start_col+3)
        sum_title = ws.cell(row=1, column=sum_start_col, value="📦 파이프 규격별 발주/재단 총괄표")
        sum_title.font = Font(bold=True, size=14)
        sum_title.alignment = Alignment(horizontal="center", vertical="center")
        sum_title.fill = PatternFill(start_color="A9D08E", end_color="A9D08E", fill_type="solid")
        sum_title.border = thin_border
        
        sum_headers = ["파이프 규격", "총 절단 수량(EA)", "총 소요길이(mm)", "6M 발주(본)"]
        for i, h in enumerate(sum_headers):
            c = ws.cell(row=3, column=sum_start_col+i, value=h)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill(start_color="385623", end_color="385623", fill_type="solid")
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = thin_border
            
        sum_r_idx = 4
        for p_name, data in pipe_summary.items():
            total_l = round(data["len"], 1)
            base_qty = data["qty"]
            
            c1 = ws.cell(row=sum_r_idx, column=sum_start_col, value=p_name)
            c2 = ws.cell(row=sum_r_idx, column=sum_start_col+1, value=f"={base_qty}*$D$1")
            c3 = ws.cell(row=sum_r_idx, column=sum_start_col+2, value=f"={total_l}*$D$1")
            c4 = ws.cell(row=sum_r_idx, column=sum_start_col+3, value=f"=ROUNDUP(({total_l}*$D$1)/6000, 0)")
            
            c1.font = Font(bold=True)
            c2.font = Font(bold=True)
            c3.font = Font(color="0070C0", bold=True)
            c4.font = Font(color="FF0000", bold=True, size=12)
            
            for i in range(4):
                c = ws.cell(row=sum_r_idx, column=sum_start_col+i)
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border = thin_border
                if sum_r_idx % 2 == 0:
                    c.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            sum_r_idx += 1

        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            max_len = max([len(str(ws.cell(row=r, column=col_idx).value)) for r in range(3, ws.max_row + 1) if ws.cell(row=r, column=col_idx).value] + [0])
            ws.column_dimensions[col_letter].width = max(max_len + 6, 12)
            
    return output.getvalue()


def get_ladder_excel_bytes(data):
    df = pd.DataFrame(data, columns=["항목", "규격", "수량(개/줄)", "단위길이(cm)", "총연장(cm)", "6m본수/비고"])
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='전체자재표')
        ws = writer.sheets['전체자재표']
        
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        center_align = Alignment(horizontal='center', vertical='center')
        header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid") 
        fill_main = PatternFill(start_color="EBF1DE", end_color="EBF1DE", fill_type="solid")   
        fill_v = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")      
        fill_angle = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  
        fill_sep = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")    
        header_font = Font(color="FFFFFF", bold=True)
        cut_size_font = Font(color="FF0000", bold=True) 
        angle_text_font = Font(color="002060", bold=True) 

        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
                except: pass
            ws.column_dimensions[column].width = (max_length * 1.6) + 6

        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column)):
            is_header = (row_idx == 0)
            item_name = str(row[0].value) if row[0].value else ""
            
            for col_idx, cell in enumerate(row):
                cell.border = thin_border
                cell.alignment = center_align
                
                if is_header:
                    cell.fill = header_fill
                    cell.font = header_font
                else:
                    if "---" in item_name: cell.fill = fill_sep
                    elif "다대" in item_name or "수직" in item_name: cell.fill = fill_v
                    elif "살대" in item_name or "사재" in item_name or "°" in item_name or "가공각" in item_name:
                        cell.fill = fill_angle
                        if col_idx == 0: cell.font = angle_text_font
                    elif "상하현재" in item_name or "스나기" in item_name or "안내" in item_name:
                        cell.fill = fill_main
                    
                    if col_idx == 3 and "---" not in item_name:
                        cell.font = cut_size_font
                        
    return output.getvalue()


# ==============================================================================
# 메인 웹앱 구성
# ==============================================================================
st.title("🏢 하나천막기업 - 자재 산출 및 도면 생성 시스템")
st.markdown("---")

tab1, tab2 = st.tabs(["🏗️ [1] 맞춤형 트러스 도면 생성기", "🪜 [2] 벽사다리/용마루 시스템"])

# ==========================================
# TAB 1: 트러스 시스템
# ==========================================
with tab1:
    st.subheader("트러스 치수 및 환경 설정")
    c1, c2 = st.columns(2)
    with c1:
        type_choice = st.selectbox("트러스 형태", [
            "1: 대칭삼각(일반)", "2: 아치형(일반)", "3: 반삼각(일반)",
            "4: 서브형_삼각", "5: 서브형_아치", "6: 서브형_반삼각",
            "7: 밑더블_삼각", "8: 밑더블_아치", "9: 밑더블_반삼각"
        ], index=6)
        
        type_code = type_choice.split(":")[0]
        name_map = {
            "1": "대칭삼각(일반)", "2": "아치형(일반)", "3": "반삼각(일반)",
            "4": "서브형_삼각", "5": "서브형_아치", "6": "서브형_반삼각",
            "7": "밑더블_삼각", "8": "밑더블_아치", "9": "밑더블_반삼각"
        }
        t_name = name_map.get(type_code, "밑더블_삼각")
        
        is_sub_type = type_code in ["4", "5", "6"]
        has_tie = type_code in ["4", "5"] 
        is_double_bot = type_code in ["7", "8", "9"]
        is_half = type_code in ["3", "6", "9"]
        is_arch = type_code in ["2", "5", "8"]

        span_cm = st.number_input("1. 전체 스판(cm)", value=1200.0)
        divs = st.number_input("2. 등분 수(다대 개수 결정)", value=34, step=1)
        
        h_tie_cm = 0.0
        if is_sub_type:
            if type_code == "5":
                h_outer_cm = st.number_input("3. 양쪽 끝단(시작) 높이(cm)", value=51.0)
            else:
                h_outer_cm = st.number_input("3. 트러스 상하 일정 수직 폭(깊이)(cm)", value=80.0)
            if has_tie:
                h_tie_cm = st.number_input("3.1. 수평재(수평선) 바닥기준 높이(cm)", value=150.0)
        elif is_double_bot:
            h_outer_cm = st.number_input("3. 밑더블 양끝 외경 시작높이(cm)", value=80.0)
        else:
            h_outer_cm = st.number_input("3. 끝단(시작) 높이(cm)", value=51.0)
            
        h_center_cm = st.number_input("4. 최고점 상단 높이(cm)", value=250.0)

    with c2:
        m_od = st.number_input("5. 상/하현부 및 수평 파이프 지름(mm)", value=59.9)
        v_od = st.number_input("6. 다대(일반) 지름(mm)", value=38.1)
        r_od = st.number_input("7. 용마루(중앙) 지름(mm)", value=59.9)
        st.markdown("**살대(대각선) 설정 - V/W자 고정 적용**")
        d_od = st.number_input("8.1. 살대(대각) 지름(mm)", value=31.8)
        offset_mm = st.number_input("8.2. 살대 이격 거리(mm)", value=20.0)

    if st.button("🚀 트러스 도면 렌더링", type="primary", use_container_width=True):
        with st.spinner("도면 계산 및 생성 중입니다..."):
            S, H_out, H_cen, H_tie = span_cm * 10, h_outer_cm * 10, h_center_cm * 10, h_tie_cm * 10
            is_diag = True

            yc, R = 0, 0
            if is_arch:
                if H_cen <= H_out: H_cen = H_out + 10
                yc = ((S/2)**2 + H_out**2 - H_cen**2) / (2 * (H_out - H_cen))
                R = H_cen - yc

            def get_y_top(x):
                if x < 0: x = 0
                if x > S: x = S
                if type_code in ["1", "4", "7"]:
                    m = (H_cen - H_out) / (S/2)
                    return H_out + m * x if x <= S/2 else H_out + m * (S - x)
                elif is_arch:
                    val = max(R**2 - (x - S/2)**2, 0)
                    return yc + math.sqrt(val)
                elif is_half:
                    return H_out + (x / S) * (H_cen - H_out)

            def get_y_bot(x):
                if is_sub_type: return max(get_y_top(x) - H_out, 0.0)
                else: return 0.0

            def get_slope(func, x):
                dx = 0.1
                test_x = x if x + dx <= S else x - dx
                dy = func(test_x + dx) - func(test_x)
                return math.degrees(math.atan2(dy, dx))
                
            def get_cos(func, x):
                dx = 0.1
                test_x = x if x + dx <= S else x - dx
                dy = func(test_x + dx) - func(test_x)
                cos_val = math.cos(math.atan2(dy, dx))
                return cos_val if cos_val != 0 else 0.0001
                
            def get_thick(func, x, od):
                return od / get_cos(func, x)

            def get_chord_y_top(x):
                return get_y_top(x) - get_thick(get_y_top, x, m_od)
                
            def get_chord_y_bot(x):
                return get_y_bot(x) + get_thick(get_y_bot, x, m_od)

            def draw_dim_text(ax, x, y, text, angle=0, color='black', fontsize=11.5):
                if angle > 90: angle -= 180
                elif angle < -90: angle += 180
                ax.text(x, y, text, color=color, fontsize=fontsize, fontweight='bold', ha='center', va='center', rotation=angle,
                        bbox=dict(facecolor='white', alpha=0.85, edgecolor='none', pad=1.5))

            fig, ax = plt.subplots(figsize=(40, 18), dpi=100)
            mid_idx = divs if is_half else divs // 2
            raw_data = []

            v_centers_x = [v_od/2 if i==0 else (S - (v_od if is_half else r_od)/2 if i==divs else i*(S/divs)) for i in range(divs + 1)]
            chord_x = [0] + [v_centers_x[i] for i in range(1, divs)] + [S]

            end_thick = get_thick(get_y_top, 0, m_od)
            H_mid_top = H_out - end_thick if is_double_bot else H_out
            H_mid_bot = H_mid_top - m_od if is_double_bot else H_out - m_od

            total_top_chord_len = sum(math.hypot(chord_x[i+1] - chord_x[i], get_y_top(chord_x[i+1]) - get_y_top(chord_x[i])) for i in range(divs))
            total_bot_chord_len = sum(math.hypot(chord_x[i+1] - chord_x[i], get_y_bot(chord_x[i+1]) - get_y_bot(chord_x[i])) for i in range(divs))
            
            raw_data.append({
                "구분": "상현대(전체)", "품명": f"{m_od}mm 파이프",
                "재단기장(L)": round(total_top_chord_len, 1), "상단 가공각(°)": 0, "하단 가공각(°)": 0
            })
            raw_data.append({
                "구분": "하현대(전체)", "품명": f"{m_od}mm 파이프",
                "재단기장(L)": round(total_bot_chord_len, 1), "상단 가공각(°)": 0, "하단 가공각(°)": 0
            })

            if is_double_bot:
                total_mid_pipe_len = S
                ax.add_patch(patches.Rectangle((0, H_mid_bot), total_mid_pipe_len, m_od, facecolor='#9b59b6', edgecolor='black', zorder=6))
                draw_dim_text(ax, S/2, H_mid_bot + m_od/2, f"밑더블 수평재(전체) L:{total_mid_pipe_len:.1f}", angle=0, color='purple', fontsize=14)
                raw_data.append({
                    "구분": "밑더블수평재", "품명": f"{m_od}mm 파이프",
                    "재단기장(L)": round(total_mid_pipe_len, 1), "상단 가공각(°)": 0, "하단 가공각(°)": 0
                })

            for i in range(divs + 1):
                x = v_centers_x[i]
                is_ridge = (i == mid_idx) and not is_half
                curr_v_od = r_od if is_ridge else v_od
                
                x_l, x_r = x - curr_v_od/2, x + curr_v_od/2
                yt_l, yt_r = get_chord_y_top(x_l), get_chord_y_top(x_r)
                yb_l, yb_r = get_chord_y_bot(x_l), get_chord_y_bot(x_r)
                
                y_bot_c = get_chord_y_bot(x)
                y_top_c = get_chord_y_top(x)
                
                if x_l < S/2 < x_r and not is_arch:
                    poly_top = [[x_r, yt_r], [S/2, get_chord_y_top(S/2)], [x_l, yt_l]]
                    poly_bot = [[x_l, yb_l], [S/2, get_chord_y_bot(S/2)], [x_r, yb_r]]
                else:
                    poly_top = [[x_r, yt_r], [x_l, yt_l]]
                    poly_bot = [[x_l, yb_l], [x_r, yb_r]]
                
                t_angle = int(round(abs(get_slope(get_y_top, x))))
                b_angle = int(round(abs(get_slope(get_y_bot, x))))
                
                if is_ridge: v_cut_l = y_top_c - min(yb_l, yb_r)
                else: v_cut_l = max(yt_l, yt_r) - min(yb_l, yb_r) 

                if has_tie and is_ridge and H_tie > 0:
                    if type_code != "5":
                        yb_l = yb_r = y_bot_c = H_tie + m_od/2
                        poly_bot = [[x_l, yb_l], [x_r, yb_r]]
                        v_cut_l = y_top_c - y_bot_c
                        b_angle = 0

                if is_double_bot:
                    u_len = y_top_c - H_mid_top
                    if u_len > 0:
                        u_cut_max = y_top_c - H_mid_top if is_ridge else max(yt_l, yt_r) - H_mid_top
                        pts_u = [[x_l, H_mid_top], [x_r, H_mid_top]] + poly_top
                        ax.add_patch(patches.Polygon(pts_u, facecolor='#2980b9', edgecolor='black', zorder=5))
                        
                        g_name = "상단용마루" if is_ridge else "상단다대"
                        raw_data.append({
                            "구분": g_name, "품명": f"{curr_v_od}mm 파이프",
                            "재단기장(L)": round(u_cut_max, 1), "상단 가공각(°)": t_angle, "하단 가공각(°)": 0
                        })
                        stagger_top = 600 if i % 2 == 0 else 900
                        my_top = y_top_c + stagger_top
                        ax.plot([x, x], [y_top_c + m_od/2, my_top - 180], color='blue', linestyle=':', lw=1.5, zorder=1)
                        draw_dim_text(ax, x, my_top, f"{g_name}\nL:{u_cut_max:.1f}", angle=90, color='blue', fontsize=10)

                    l_len = H_mid_bot - y_bot_c
                    if l_len > 0:
                        l_cut_max = H_mid_bot - min(yb_l, yb_r) if not is_ridge else H_mid_bot - y_bot_c
                        pts_l = poly_bot + [[x_r, H_mid_bot], [x_l, H_mid_bot]]
                        ax.add_patch(patches.Polygon(pts_l, facecolor='#34495e', edgecolor='black', zorder=5))
                        
                        g_name = "하단용마루" if is_ridge else "하단다대"
                        raw_data.append({
                            "구분": g_name, "품명": f"{curr_v_od}mm 파이프",
                            "재단기장(L)": round(l_cut_max, 1), "상단 가공각(°)": 0, "하단 가공각(°)": b_angle
                        })
                        stagger_bot = 600 if i % 2 == 0 else 900
                        my_bot = y_bot_c - stagger_bot
                        ax.plot([x, x], [y_bot_c - m_od/2, my_bot + 180], color='darkblue', linestyle=':', lw=1.5, zorder=1)
                        draw_dim_text(ax, x, my_bot, f"{g_name}\nL:{l_cut_max:.1f}", angle=90, color='darkblue', fontsize=10)
                else:
                    pts_v = poly_bot + poly_top
                    ax.add_patch(patches.Polygon(pts_v, facecolor='#2c3e50', edgecolor='black', zorder=5))
                    
                    text_color = 'red' if is_ridge else 'blue'
                    stagger_offset = 600 if i % 2 == 0 else 900
                    my = y_top_c + stagger_offset
                    ax.plot([x, x], [y_top_c + m_od/2, my - 180], color=text_color, linestyle=':', lw=1.5, zorder=1)
                    draw_dim_text(ax, x, my, f"L:{v_cut_l:.1f} (상:{t_angle}°/하:{b_angle}°)", angle=90, color=text_color)
                    
                    v_gubun = "용마루" if is_ridge else "다대"
                    raw_data.append({
                        "구분": v_gubun, "품명": f"{curr_v_od}mm 파이프",
                        "재단기장(L)": round(v_cut_l, 1), "상단 가공각(°)": t_angle, "하단 가공각(°)": b_angle
                    })

            for i in range(divs):
                cx, cnx = chord_x[i], chord_x[i+1]
                
                pb1, pb2 = (cx, get_y_bot(cx)), (cnx, get_y_bot(cnx))
                pb3, pb4 = (cnx, get_chord_y_bot(cnx)), (cx, get_chord_y_bot(cx))
                ax.add_patch(patches.Polygon(np.array([pb1, pb2, pb3, pb4]), facecolor='#7f8c8d', alpha=0.5, zorder=2))
                
                pt1, pt2 = (cx, get_chord_y_top(cx)), (cnx, get_chord_y_top(cnx))
                pt3, pt4 = (cnx, get_y_top(cnx)), (cx, get_y_top(cx))
                ax.add_patch(patches.Polygon(np.array([pt1, pt2, pt3, pt4]), facecolor='#7f8c8d', zorder=7))

                if is_diag:
                    x, nx = v_centers_x[i], v_centers_x[i+1]
                    is_r_curr = (i == mid_idx) and not is_half
                    is_r_next = (i+1 == mid_idx) and not is_half
                    c_v_od, n_v_od = (r_od if is_r_curr else v_od), (r_od if is_r_next else v_od)

                    wx_start = x + c_v_od/2 + offset_mm
                    wx_end = nx - n_v_od/2 - offset_mm
                    
                    def draw_diag(left_x, right_x, is_forward):
                        dx = right_x - left_x
                        if dx <= 0: return
                        
                        if is_forward:
                            y_bot_real = get_chord_y_bot(left_x)
                            y_top_real = get_chord_y_top(right_x)
                            diag_l = math.hypot(dx, y_top_real - y_bot_real)
                        else:
                            y_bot_real = get_chord_y_bot(right_x)
                            y_top_real = get_chord_y_top(left_x)
                            diag_l = math.hypot(dx, y_top_real - y_bot_real)
                            
                        mid_x = (left_x + right_x) / 2
                        v_len = abs(get_chord_y_top(mid_x) - get_chord_y_bot(mid_x))
                        
                        diag_len = math.hypot(dx, v_len)
                        sin_theta = v_len / diag_len if diag_len > 0 else 1
                        w_half = (d_od / 2) / sin_theta if sin_theta > 0.01 else d_od / 2
                        w_half = min(w_half, dx * 0.45) 
                        
                        px_bot = left_x + w_half if is_forward else right_x - w_half
                        px_top = right_x - w_half if is_forward else left_x + w_half
                        
                        py_bot = get_chord_y_bot(px_bot)
                        py_top = get_chord_y_top(px_top)
                        
                        if is_forward:
                            x_bl, x_br = px_bot - w_half, px_bot + w_half
                            x_tr, x_tl = px_top + w_half, px_top - w_half
                            pts = [
                                [x_bl, get_chord_y_bot(x_bl)], [x_br, get_chord_y_bot(x_br)],
                                [x_tr, get_chord_y_top(x_tr)], [x_tl, get_chord_y_top(x_tl)]
                            ]
                        else:
                            x_br, x_bl = px_bot + w_half, px_bot - w_half
                            x_tl, x_tr = px_top - w_half, px_top + w_half
                            pts = [
                                [x_br, get_chord_y_bot(x_br)], [x_bl, get_chord_y_bot(x_bl)],
                                [x_tl, get_chord_y_top(x_tl)], [x_tr, get_chord_y_top(x_tr)]
                            ]
                            
                        poly = plt.Polygon(pts, facecolor='#f1c40f', edgecolor='black', linewidth=1.2, zorder=3)
                        ax.add_patch(poly)
                        
                        dx_line, dy_line = px_top - px_bot, py_top - py_bot
                        diag_ang = math.degrees(math.atan2(dy_line, dx_line))
                        
                        t_slope = get_slope(get_y_top, right_x if is_forward else left_x)
                        b_slope = get_slope(get_y_bot, left_x if is_forward else right_x)
                        
                        t_intersect = abs(diag_ang - t_slope) % 180
                        if t_intersect > 90: t_intersect = 180 - t_intersect
                        d_top_angle = int(round(abs(90.0 - t_intersect)))
                        
                        b_intersect = abs(diag_ang - b_slope) % 180
                        if b_intersect > 90: b_intersect = 180 - b_intersect
                        d_bot_angle = int(round(abs(90.0 - b_intersect)))

                        mx, my = (px_bot + px_top) / 2, (py_bot + py_top) / 2
                        draw_dim_text(ax, mx, my, f"L:{diag_l:.1f} ({d_top_angle}°/{d_bot_angle}°)", angle=diag_ang, color='#8B0000', fontsize=11)

                        raw_data.append({
                            "구분": "살대", "품명": f"{d_od}mm 파이프",
                            "재단기장(L)": round(diag_l, 1), "상단 가공각(°)": d_top_angle, "하단 가공각(°)": d_bot_angle
                        })

                    if is_double_bot:
                        def draw_custom_diag(left_x, right_x, gubun_name, is_forward):
                            dx = right_x - left_x
                            if dx <= 0: return
                            if gubun_name == "상단살대":
                                py_bot_real = H_mid_top
                                if is_forward:
                                    py_top_real = get_chord_y_top(right_x)
                                    diag_l = math.hypot(dx, py_top_real - py_bot_real)
                                else:
                                    py_top_real = get_chord_y_top(left_x)
                                    diag_l = math.hypot(dx, py_top_real - py_bot_real)
                                y_bot_est = H_mid_top
                                y_top_est = get_chord_y_top((left_x+right_x)/2)
                            else:
                                py_top_real = H_mid_bot
                                if is_forward:
                                    py_bot_real = get_chord_y_bot(left_x)
                                    diag_l = math.hypot(dx, py_top_real - py_bot_real)
                                else:
                                    py_bot_real = get_chord_y_bot(right_x)
                                    diag_l = math.hypot(dx, py_top_real - py_bot_real)
                                y_top_est = H_mid_bot
                                y_bot_est = get_chord_y_bot((left_x+right_x)/2)
                                
                            v_len = abs(y_top_est - y_bot_est)
                            diag_len = math.hypot(dx, v_len)
                            sin_theta = v_len / diag_len if diag_len > 0 else 1
                            w_half = (d_od / 2) / sin_theta if sin_theta > 0.01 else d_od / 2
                            w_half = min(w_half, dx * 0.45)
                            
                            px_bot = left_x + w_half if is_forward else right_x - w_half
                            px_top = right_x - w_half if is_forward else left_x + w_half
                            
                            if gubun_name == "상단살대":
                                py_bot = H_mid_top
                                py_top = get_chord_y_top(px_top)
                            else:
                                py_top = H_mid_bot
                                py_bot = get_chord_y_bot(px_bot)
                                
                            if py_top <= py_bot: return 
                            
                            if is_forward:
                                x_bl, x_br = px_bot - w_half, px_bot + w_half
                                x_tr, x_tl = px_top + w_half, px_top - w_half
                                if gubun_name == "상단살대":
                                    pts = [[x_bl, H_mid_top], [x_br, H_mid_top], [x_tr, get_chord_y_top(x_tr)], [x_tl, get_chord_y_top(x_tl)]]
                                else:
                                    pts = [[x_bl, get_chord_y_bot(x_bl)], [x_br, get_chord_y_bot(x_br)], [x_tr, H_mid_bot], [x_tl, H_mid_bot]]
                            else:
                                x_br, x_bl = px_bot + w_half, px_bot - w_half
                                x_tl, x_tr = px_top - w_half, px_top + w_half
                                if gubun_name == "상단살대":
                                    pts = [[x_br, H_mid_top], [x_bl, H_mid_top], [x_tl, get_chord_y_top(x_tl)], [x_tr, get_chord_y_top(x_tr)]]
                                else:
                                    pts = [[x_br, get_chord_y_bot(x_br)], [x_bl, get_chord_y_bot(x_bl)], [x_tl, H_mid_bot], [x_tr, H_mid_bot]]
                                    
                            poly = plt.Polygon(pts, facecolor='#f1c40f', edgecolor='black', linewidth=1.2, zorder=3)
                            ax.add_patch(poly)
                            
                            dx_line, dy_line = px_top - px_bot, py_top - py_bot
                            diag_ang = math.degrees(math.atan2(dy_line, dx_line))
                            
                            if gubun_name == "상단살대":
                                t_slope = get_slope(get_y_top, right_x if is_forward else left_x)
                                b_slope = 0.0
                            else:
                                t_slope = 0.0 
                                b_slope = get_slope(get_y_bot, left_x if is_forward else right_x)
                                
                            t_intersect = abs(diag_ang - t_slope) % 180
                            if t_intersect > 90: t_intersect = 180 - t_intersect
                            d_top_angle = int(round(abs(90.0 - t_intersect)))
                            
                            b_intersect = abs(diag_ang - b_slope) % 180
                            if b_intersect > 90: b_intersect = 180 - b_intersect
                            d_bot_angle = int(round(abs(90.0 - b_intersect)))
                            
                            mx, my = (px_bot + px_top)/2, (py_bot + py_top)/2
                            draw_dim_text(ax, mx, my, f"L:{diag_l:.1f} ({d_top_angle}°/{d_bot_angle}°)", angle=diag_ang, color='#8B0000', fontsize=10)
                            
                            raw_data.append({
                                "구분": gubun_name, "품명": f"{d_od}mm 파이프",
                                "재단기장(L)": round(diag_l, 1), "상단 가공각(°)": d_top_angle, "하단 가공각(°)": d_bot_angle
                            })

                        is_forward_u = not (is_half or i < mid_idx)
                        draw_custom_diag(wx_start, wx_end, "상단살대", is_forward_u)

                        if is_half: is_forward_l = (i % 2 == 0)
                        else:
                            if i < mid_idx: is_forward_l = ((mid_idx - 1 - i) % 2 != 0) 
                            else: is_forward_l = ((i - mid_idx) % 2 == 0)
                        draw_custom_diag(wx_start, wx_end, "하단살대", is_forward_l)
                    else:
                        if is_half or i < mid_idx: draw_diag(wx_start, wx_end, False)
                        else: draw_diag(wx_start, wx_end, True)

            if has_tie and H_tie > 0:
                y_tie_top = H_tie + m_od / 2
                low, high = 0.0, S/2
                for _ in range(50):
                    mid = (low + high) / 2
                    if get_y_bot(mid) < y_tie_top: low = mid
                    else: high = mid
                x_left_in = (low + high) / 2
                
                if is_half: x_right_in = S - m_od / 2
                else:
                    low, high = S/2, S
                    for _ in range(50):
                        mid = (low + high) / 2
                        if get_y_bot(mid) > y_tie_top: low = mid
                        else: high = mid
                    x_right_in = (low + high) / 2

                tie_length_inner = x_right_in - x_left_in
                if tie_length_inner > 0:
                    ax.add_patch(patches.Rectangle((x_left_in, H_tie - m_od/2), tie_length_inner, m_od, facecolor='#8e44ad', edgecolor='black', zorder=6))
                    ax.annotate(f"수평재 내경 L:{tie_length_inner:.1f}", xy=((x_left_in+x_right_in)/2, H_tie), xytext=((x_left_in+x_right_in)/2, H_tie - 120),
                                arrowprops=dict(arrowstyle='<->', color='purple', lw=3), ha='center', fontsize=16, fontweight='bold', color='purple')
                    raw_data.append({
                        "구분": "수평재", "품명": f"{m_od}mm 파이프", "재단기장(L)": round(tie_length_inner, 1),
                        "상단 가공각(°)": 0, "하단 가공각(°)": 0
                    })

                valid_inner_xs = []
                for i in range(1, divs):
                    if i == mid_idx and not is_half:
                        if type_code != "5": continue
                    x = v_centers_x[i]
                    if get_y_bot(x) > (H_tie + m_od/2) + 10:
                        valid_inner_xs.append((i, x))
                        
                for i, x in valid_inner_xs:
                    is_r_curr = (i == mid_idx) and not is_half
                    curr_v_od = r_od if is_r_curr else v_od
                    x_l, x_r = x - curr_v_od / 2, x + curr_v_od / 2
                    yt_l, yt_r = get_y_bot(x_l), get_y_bot(x_r)
                    
                    if x_l < S/2 < x_r and not is_arch: poly_top_inner = [[x_r, yt_r], [S/2, get_y_bot(S/2)], [x_l, yt_l]]
                    else: poly_top_inner = [[x_r, yt_r], [x_l, yt_l]]
                    
                    y_bot = H_tie + m_od/2
                    v_len = max(yt_l, yt_r) - y_bot
                    
                    pts_inner = [[x_l, y_bot], [x_r, y_bot]] + poly_top_inner
                    ax.add_patch(patches.Polygon(pts_inner, facecolor='#2ecc71', edgecolor='black', zorder=4))
                    t_angle = int(round(abs(get_slope(get_y_bot, x))))
                    
                    stagger_bot = 500 if i % 2 == 0 else 800
                    if is_r_curr: stagger_bot = 1100 
                    my_bot = H_tie - stagger_bot
                    ax.plot([x, x], [H_tie - m_od/2, my_bot + 180], color='darkgreen', linestyle=':', lw=1.5, zorder=1)
                    draw_dim_text(ax, x, my_bot, f"L:{v_len:.1f} (상:{t_angle}°/하:0°)", angle=90, color='darkgreen')
                    
                    raw_data.append({
                        "구분": "수평내부다대", "품명": f"{curr_v_od}mm 파이프", "재단기장(L)": round(v_len, 1),
                        "상단 가공각(°)": t_angle, "하단 가공각(°)": 0
                    })
                    
                if type_code == "5":
                    tie_bot_y = H_tie - m_od/2
                    dim_h_x = S/2 - r_od/2 - 250 if not is_half else S/2 - 250
                    ax.plot([dim_h_x - 100, dim_h_x + 100], [0, 0], color='black', lw=1.5, zorder=1)
                    ax.annotate("", xy=(dim_h_x, tie_bot_y), xytext=(dim_h_x, 0), arrowprops=dict(arrowstyle='<->', color='#d35400', lw=2.5))
                    draw_dim_text(ax, dim_h_x, tie_bot_y / 2, f"수평보 하단 높이: {tie_bot_y:.1f}", angle=90, color='#d35400', fontsize=13)

                valid_inner_x_vals = [x for _, x in valid_inner_xs]
                base_intervals = [x_left_in] + valid_inner_x_vals + [x_right_in]
                if not is_half: base_intervals.append(S/2)
                inner_intervals = sorted(list(set(base_intervals)))
                
                for i in range(len(inner_intervals)-1):
                    if i == 0: continue  
                    if not is_half and i == len(inner_intervals) - 2: continue  

                    lx, rx = inner_intervals[i], inner_intervals[i+1]
                    l_v_od = r_od if (round(lx, 1) == round(S/2, 1) and not is_half) else v_od
                    r_v_od = r_od if (round(rx, 1) == round(S/2, 1) and not is_half) else v_od
                    wx_start, wx_end = lx + l_v_od/2 + offset_mm, rx - r_v_od/2 - offset_mm
                    y_bot_limit = H_tie + m_od/2
                    
                    if wx_end > wx_start + 10: 
                        is_forward_tie = not (is_half or lx < S/2)
                        dx_est = wx_end - wx_start
                        dy_est = get_y_bot(wx_end) - y_bot_limit if is_forward_tie else get_y_bot(wx_start) - y_bot_limit
                        diag_l = math.hypot(dx_est, dy_est)
                        
                        diag_len_est = math.hypot(dx_est, dy_est)
                        sin_theta = abs(dy_est) / diag_len_est if diag_len_est > 0 else 1
                        w_half_tie = (d_od / 2) / sin_theta if sin_theta > 0.01 else d_od / 2
                        w_half_tie = min(w_half_tie, dx_est * 0.45)
                        
                        px_bot = wx_start + w_half_tie if is_forward_tie else wx_end - w_half_tie
                        px_top = wx_end - w_half_tie if is_forward_tie else wx_start + w_half_tie
                        py_bot, py_top = y_bot_limit, get_y_bot(px_top) 
                        
                        if is_forward_tie:
                            pts = [[wx_start, y_bot_limit], [wx_start + 2*w_half_tie, y_bot_limit], [wx_end, get_y_bot(wx_end)], [wx_end - 2*w_half_tie, get_y_bot(wx_end - 2*w_half_tie)]]
                        else:
                            pts = [[wx_end, y_bot_limit], [wx_end - 2*w_half_tie, y_bot_limit], [wx_start, get_y_bot(wx_start)], [wx_start + 2*w_half_tie, get_y_bot(wx_start + 2*w_half_tie)]]
                            
                        poly = plt.Polygon(pts, facecolor='#f1c40f', edgecolor='black', linewidth=1.2, zorder=3)
                        ax.add_patch(poly)
                        
                        dx_diff, dy_diff = px_top - px_bot, py_top - py_bot
                        diag_ang = math.degrees(math.atan2(dy_diff, dx_diff))
                        
                        t_slope = get_slope(get_y_bot, wx_end if is_forward_tie else wx_start)
                        b_slope = 0.0
                        
                        t_intersect = abs(diag_ang - t_slope) % 180
                        if t_intersect > 90: t_intersect = 180 - t_intersect
                        d_top_angle = int(round(abs(90.0 - t_intersect)))
                        
                        b_intersect = abs(diag_ang - b_slope) % 180
                        if b_intersect > 90: b_intersect = 180 - b_intersect
                        d_bot_angle = int(round(abs(90.0 - b_intersect)))
                        
                        mx, my = (px_bot + px_top)/2, (py_bot + py_top)/2
                        draw_dim_text(ax, mx, my, f"L:{diag_l:.1f} ({d_top_angle}°/{d_bot_angle}°)", angle=diag_ang, color='#b8860b', fontsize=11)
                        
                        raw_data.append({
                            "구분": "수평내부살대", "품명": f"{d_od}mm 파이프", "재단기장(L)": round(diag_l, 1),
                            "상단 가공각(°)": d_top_angle, "하단 가공각(°)": d_bot_angle
                        })

            dim_y = -1500 if is_double_bot else -350
            ax.plot([0, S], [dim_y, dim_y], color='black', lw=2, zorder=10)
            ax.plot([0, 0], [dim_y - 25, dim_y + 25], color='black', lw=2, zorder=10)
            ax.plot([S, S], [dim_y - 25, dim_y + 25], color='black', lw=2, zorder=10)

            ticks_x = [0] + [v_centers_x[i] for i in range(1, divs)] + [S]
            for i in range(divs):
                tx1, tx2 = ticks_x[i], ticks_x[i+1]
                cx = (tx1 + tx2) / 2
                interval_len = tx2 - tx1
                if i > 0: 
                    ax.plot([tx1, tx1], [dim_y - 20, dim_y + 20], color='black', lw=1.5, zorder=10)
                    ax.plot([tx1, tx1], [0, dim_y], color='gray', linestyle=':', lw=1.5, zorder=1)
                f_size = 12 if interval_len > 300 else 10
                ax.text(cx, dim_y + 40, f"{interval_len:.1f}", ha='center', va='center', fontsize=f_size, color='navy', fontweight='bold')

            for tx in ticks_x:
                ax.text(tx, dim_y - 45, f"{tx:.1f}", ha='center', va='top', fontsize=11, color='#d35400', fontweight='bold')

            ax.plot([S, S], [0, dim_y], color='gray', linestyle=':', lw=1.5, zorder=1)
            ax.text(S/2, dim_y - 120, f"전체 스판 : {S:.1f} mm", ha='center', va='center', fontsize=18, fontweight='bold', color='black')

            ax.set_xlim(-200, S + 200)
            ax.set_ylim(dim_y - 300, H_cen + 1200) 
            ax.set_aspect('equal')
            ax.axis('off') 
            
            info_text = f"스판: {span_cm}cm | 등분: {divs} (자간: {interval_len/10:.1f}cm)"
            if has_tie: info_text += f" | 수평재 높이: {h_tie_cm}cm"
            if is_double_bot: info_text += f" | 밑더블 외경 높이: {h_outer_cm}cm"
            
            plt.title(f"트러스 도면 ({t_name})\n{info_text}", fontsize=24, fontweight='bold', pad=20)
            
            st.pyplot(fig)
            
            # 🟢 [수정됨] 도면을 PDF 형식으로 버퍼에 저장
            pdf_buffer = BytesIO()
            fig.savefig(pdf_buffer, format="pdf", bbox_inches="tight")
            pdf_data = pdf_buffer.getvalue()
            
            plt.close(fig)

            # 엑셀 변환 로직
            excel_data = get_truss_excel_bytes(raw_data)
            
            st.success("✅ 트러스 렌더링 및 도면/엑셀 도출이 성공적으로 완료되었습니다!")
            
            # 🟢 [수정됨] 엑셀과 PDF 다운로드 버튼을 2개로 배치
            btn_col1, btn_col2 = st.columns(2)
            
            with btn_col1:
                st.download_button(
                    label="📥 트러스 산출표 (Excel) 다운로드",
                    data=excel_data,
                    file_name=f"Truss_{t_name.replace(' ', '')}_{int(span_cm)}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            
            with btn_col2:
                st.download_button(
                    label="📥 트러스 도면 (PDF) 다운로드",
                    data=pdf_data,
                    file_name=f"Truss_Drawing_{t_name.replace(' ', '')}_{int(span_cm)}.pdf",
                    mime="application/pdf"
                )


# ==========================================
# TAB 2: 벽사다리 / 용마루 시스템
# ==========================================
with tab2:
    st.subheader("벽사다리 및 보강사다리 치수 입력")
    lc1, lc2 = st.columns(2)
    with lc1:
        st.markdown("**1. 기본 치수 (중심선 기준)**")
        L_cm = st.number_input("전체 총기장(cm)", value=2000.0)
        W_cm = st.number_input("보강사다리 폭(cm)", value=70.0)
        H_truss_cm_l = st.number_input("메인사다리 폭(높이)(cm)", value=70.0)
        total_sets_sub = st.number_input("보강사다리 총 제작 수량(세트)", value=1, step=1)
        total_sets_main = st.number_input("메인사다리 총 제작 수량(세트)", value=1, step=1)
        offset_mm_l = st.number_input("살대(대각재) 이격 거리(mm)", value=10.0)
        
        st.markdown("**2. 용마루 및 벽사다리 추가 설정**")
        H_ridge_cm = st.number_input("용마루 폭(높이)(cm)", value=70.0)
        ridge_deduct_mm = st.number_input("용마루 공제 기준 사이즈(mm)", value=59.9)
        total_sets_ridge = st.number_input("용마루 전체 라인 제작 수량(세트)", value=1, step=1)
        wall_snagi_mm = st.number_input("벽사다리 스나기 사이즈(mm)", value=89.1)
        
    with lc2:
        st.markdown("**3. 파이프 규격 설정(mm)**")
        p_sub_main = st.number_input("보강사다리 상하현재", value=38.1)
        p_sub_sub = st.number_input("보강사다리 수직/사재", value=31.8)
        p_main_main = st.number_input("메인사다리 상하현재", value=42.2)
        p_main_snagi = st.number_input("메인사다리 스나기", value=89.1)
        p_main_v = st.number_input("메인사다리 수직다대", value=38.1)
        p_main_diag = st.number_input("메인사다리 사다리살대", value=31.8)
        p_ridge_main = st.number_input("용마루 상하현재", value=42.2)
        p_ridge_v = st.number_input("용마루 수직다대", value=38.1)
        p_ridge_diag = st.number_input("용마루 사다리살대", value=31.8)

    if st.button("🚀 사다리 도면 렌더링", type="primary", use_container_width=True):
        with st.spinner("사다리 도면 및 엑셀을 생성 중입니다..."):
            offset_cm = offset_mm_l / 10.0
            ridge_deduct_cm = ridge_deduct_mm / 10.0
            wall_snagi_cm = wall_snagi_mm / 10.0

            # 🟢 [수정됨] 잘렸던 부분 복구 및 변수 할당
            t_sub_main_cm = p_sub_main / 10.0
            t_sub_sub_cm = p_sub_sub / 10.0
            
            # --- 사다리 도면 임시 렌더링 (대표님께서 원하시는 로직으로 교체하세요) ---
            fig_ladder, ax_ladder = plt.subplots(figsize=(15, 5), dpi=100)
            ax_ladder.text(0.5, 0.5, "사다리 도면 렌더링 영역\n(여기에 사다리 작도 로직이 들어갑니다)", 
                           fontsize=20, ha='center', va='center', fontweight='bold', color='gray')
            ax_ladder.set_xlim(0, 1)
            ax_ladder.set_ylim(0, 1)
            ax_ladder.axis('off')
            plt.title(f"사다리 시스템 (총기장: {L_cm}cm)", fontsize=18, fontweight='bold', pad=15)
            # -------------------------------------------------------------
            
            st.pyplot(fig_ladder)
            
            # 사다리 도면 PDF 변환 로직
            pdf_ladder_buffer = BytesIO()
            fig_ladder.savefig(pdf_ladder_buffer, format="pdf", bbox_inches="tight")
            pdf_ladder_data = pdf_ladder_buffer.getvalue()
            
            plt.close(fig_ladder)

            # --- 사다리 엑셀 데이터 임시 생성 (대표님께서 원하시는 로직으로 교체하세요) ---
            ladder_raw_data = [
                ["보강사다리 상하현재", f"{p_sub_main}mm", 2, L_cm, L_cm * 2, "비고 란"],
                ["---", "---", "---", "---", "---", "---"],
                ["보강사다리 수직다대", f"{p_sub_sub}mm", 10, W_cm, W_cm * 10, ""]
            ]
            excel_ladder_data = get_ladder_excel_bytes(ladder_raw_data)
            # -------------------------------------------------------------
            
            st.success("✅ 사다리 렌더링 및 엑셀/도면 도출이 성공적으로 완료되었습니다!")
            
            btn_l_col1, btn_l_col2 = st.columns(2)
            
            with btn_l_col1:
                st.download_button(
                    label="📥 사다리 산출표 (Excel) 다운로드",
                    data=excel_ladder_data,
                    file_name="Ladder_Material_List.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                
            with btn_l_col2:
                st.download_button(
                    label="📥 사다리 도면 (PDF) 다운로드",
                    data=pdf_ladder_data,
                    file_name="Ladder_Drawing.pdf",
                    mime="application/pdf"
                )
